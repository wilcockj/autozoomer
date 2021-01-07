import time
import webbrowser
import schedule
import pyautogui
import pathlib
import os
from openpyxl import load_workbook
import re
import logging
import requests
import configparser
import cv2
import random
from datetime import datetime
from subprocess import Popen, PIPE
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters
from sys import platform
import pyscreeze
import subprocess

chat_id = ""
api_key = ""

# Enable logging
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)
# logging.disable()
parent = pathlib.Path(__file__).resolve().parent
mypath = pathlib.Path(parent)
config = configparser.ConfigParser()

# TODO
# getalltitles etc. does not work with linux must find alternative or remove functionality
# add telegram bot on another thread to exit gracefully
# need to fix if excel row is empty
# add functionality for consenting to being recorded
# implement Start Date and End Date only join meetings during that interval
# add functionality to connect a prerecorded video to the meeting
# add abbreviations for days in /sch {day} command

# add option to make bot fully autonomous, needs to click button to join meeting that it is already in or leave meetings once they are finished
def authenticator(func):
    def helper(x, y):
        if str(x.message.chat_id) == str(chat_id):
            func(x, y)
        else:
            logging.info(
                f"Unauthenticated user tried to send command, chat id ={x.message.chat_id}"
            )
            update.message.reply_text("Unauthenticated user")

    return helper


class ZoomBot:
    def __init__(self):
        self.message = ""
        self.meetingdata = self.loadexcelfile()
        self.meetingdays = self.meetingarray(7)
        self.days = [
            "MONDAY",
            "TUESDAY",
            "WEDNESDAY",
            "THURSDAY",
            "FRIDAY",
            "SATURDAY",
            "SUNDAY",
        ]
        self.sendlinebreaks()
        self.createschedule()

    def meetingarray(self, num):
        cols = num
        arr = []
        for i in range(cols):
            arr.append([])
        return arr

    def sendlinebreaks(self):
        newlines = ""
        for x in range(40):
            newlines += "." + "\n"
        sendmessage(newlines)

    def sendinfo(self, unused_bot, context):
        if len(context.args) > 0 and str(unused_bot.message.chat_id) == str(chat_id):
            daydata = self.getdaynum(context.args[0])
            if daydata[0] != -1:
                dayssched = f"Your schedule for {daydata[1]} is:\n\n"
                for x in range(len(self.meetingdays[daydata[0]])):
                    dayssched += f"{self.meetingdays[daydata[0]][x]}" + "\n\n"
                if dayssched != f"Your schedule for {daydata[1]} is:\n\n":
                    sendmessage(dayssched)
                else:
                    sendmessage(f"No Classes on {daydata[1]}!")
            else:
                sendmessage(
                    f"{context.args[0]} is an invalid day \nYou can use {' '.join(map(lambda x:x.capitalize(),self.days))}"
                )
        else:
            sendmessage(self.message)

    def getdaynum(self, day):
        day = day.upper()
        days = [
            ["MONDAY", "MON", "M"],
            ["TUESDAY", "TUES", "TU"],
            ["WEDNESDAY", "WEDS", "WED", "W"],
            ["THURSDAY", "THURS"],
            ["FRIDAY", "FRI", "F"],
            ["SATURDAY", "SAT"],
            ["SUNDAY", "SUN", "SU"],
        ]
        for x in range(len(days)):
            if any(day in word for word in days[x]):
                return [x, days[x][0].capitalize()]
        return [-1, -1]

    def loadexcelfile(self):
        excelpath = mypath / "docs" / "schedule.xlsx"
        wb = load_workbook(excelpath)
        sheet = wb["Sheet1"]
        numofcols = sheet.max_column
        numofrows = sheet.max_row
        zoomdata = []
        loopdata = []
        for x in range(2, numofrows + 1):

            if sheet.cell(row=x, column=1).value is not None:
                zoomdata.append([])
            for y in range(1, numofcols + 1):
                cellvalue = sheet.cell(row=x, column=y).value
                if y == 1 and cellvalue is None:
                    break
                else:
                    zoomdata[len(zoomdata) - 1].append(cellvalue)
        return zoomdata

    def timefixer(self, mytime):
        splittime = mytime.split(":")
        time = f"{splittime[0]}:{splittime[1]}"
        t = datetime.strptime(time, "%H:%M")
        timevalue_12hour = t.strftime("%I:%M %p")
        return timevalue_12hour

    def createschedule(self):
        # [1] is link
        # [2] is password, prob not needed many classes have code included
        # [3] is time
        # [4-10] monday - sunday
        zoomlinks = []
        zoompasses = []
        zoomtimes = []
        meetingnames = []
        dayslist = []
        for x in range(len(self.meetingdata)):
            meetingnames.append(self.meetingdata[x][0])
            dayslist.append([])
            zoomlinks.append(self.meetingdata[x][1].strip())
            if self.meetingdata[x][2] is not None:
                zoompasses.append(self.meetingdata[x][2])
            else:
                zoompasses.append(-1)
                # if no password we have only link then regex for pass and code to make a better link
                # p = re.compile("(\d{11}).*pwd=(.*)&")
                p = re.compile("(\d{10,11}).*pwd=(.{32})")
                m = p.search(zoomlinks[x])
                if m:
                    # creates a direct join link to zoom application
                    # another option is to be able to forgo application entirely with
                    # link this https://www.zoom.us/wc/join/94165984842?pwd=ME1OemMrdHdUSElGaXdobkN4Z2NzQT09
                    # which opens browser version of zoom
                    # could have an option but might be confusing to many
                    if m.group(1) and m.group(2):
                        zoomlinks[x] = (
                            "zoommtg://www.zoom.us/join?action=join&confno="
                            + m.group(1)
                            + "&pwd="
                            + m.group(2)
                        )
                    # zoomlinks[x] = "https://www.zoom.us/wc/join/" + m.group(1) + "?pwd=" + m.group(2)

            zoomtimes.append(str(self.meetingdata[x][3]))
            for i in range(4, 11, 1):
                if self.meetingdata[x][i] is not None:
                    dayslist[x].append(self.getday(i))
        schedule_message = ""
        for x in range(len(self.meetingdata)):
            for i in range(len(dayslist[x])):
                self.setschedule(
                    dayslist[x][i],
                    zoomtimes[x],
                    [zoomlinks[x], zoompasses[x], meetingnames[x]],
                )
            timevalue_12hour = self.timefixer(zoomtimes[x])
            schedule_message += (
                f"Scheduling {meetingnames[x].upper()} meeting on {', '.join([x.capitalize() for x in dayslist[x]])} joining at {str(timevalue_12hour)} "
                + "\n\n"
            )
        logging.debug(zoomlinks)
        sendmessage(schedule_message)
        self.message = schedule_message

    def getday(self, daynum):
        return self.days[daynum - 4]

    def setschedule(self, day, time, zoomdata):
        splittime = time.split(":")
        time = f"{splittime[0]}:{splittime[1]}"
        if day.upper() == "MONDAY":
            print(
                f"Setting schedule for {zoomdata[2].upper()} on {day.capitalize()} joining conference at {str(time)}"
            )
            schedule.every().monday.at(str(time)).do(joinzoommeeting, zoomdata)
            self.meetingdays[0].append(
                f"{zoomdata[2].upper()} at {str(self.timefixer(time))}"
            )
        elif day.upper() == "TUESDAY":
            print(
                f"Setting schedule for {zoomdata[2].upper()} on {day.capitalize()} joining conference at {str(time)}"
            )
            schedule.every().tuesday.at(str(time)).do(joinzoommeeting, zoomdata)
            self.meetingdays[1].append(
                f"{zoomdata[2].upper()} at {str(self.timefixer(time))}"
            )
        elif day.upper() == "WEDNESDAY":
            print(
                f"Setting schedule for {zoomdata[2].upper()} on {day.capitalize()} joining conference at {str(time)}"
            )
            schedule.every().wednesday.at(str(time)).do(joinzoommeeting, zoomdata)
            self.meetingdays[2].append(
                f"{zoomdata[2].upper()} at {str(self.timefixer(time))}"
            )
        elif day.upper() == "THURSDAY":
            print(
                f"Setting schedule for {zoomdata[2].upper()} on {day.capitalize()} joining conference at {str(time)}"
            )
            schedule.every().thursday.at(str(time)).do(joinzoommeeting, zoomdata)
            self.meetingdays[3].append(
                f"{zoomdata[2].upper()} at {str(self.timefixer(time))}"
            )
        elif day.upper() == "FRIDAY":
            print(
                f"Setting schedule for {zoomdata[2].upper()} on {day.capitalize()} joining conference at {str(time)}"
            )
            schedule.every().friday.at(str(time)).do(joinzoommeeting, zoomdata)
            self.meetingdays[4].append(
                f"{zoomdata[2].upper()} at {str(self.timefixer(time))}"
            )
        elif day.upper() == "SATURDAY":
            print(
                f"Setting schedule for {zoomdata[2].upper()} on {day.capitalize()} joining conference at {str(time)}"
            )
            schedule.every().saturday.at(str(time)).do(joinzoommeeting, zoomdata)
            self.meetingdays[5].append(
                f"{zoomdata[2].upper()} at {str(self.timefixer(time))}"
            )
        elif day.upper() == "SUNDAY":
            print(
                f"Setting schedule for {zoomdata[2].upper()} on {day.capitalize()} joining conference at {str(time)}"
            )
            schedule.every().sunday.at(str(time)).do(joinzoommeeting, zoomdata)
            self.meetingdays[6].append(
                f"{zoomdata[2].upper()} at {str(self.timefixer(time))}"
            )


def joinzoommeeting(info):
    # info[0] classcode info [1] password if there is one
    # print(info[0],info[1])
    # [10:] gets part after zoommtg:// to get the clickable link
    # need to send just password and conference number

    sendmessage(f"Trying to join meeting link: {info[0][10:]}")
    p = re.compile("(\d{10,11}).*pwd=(.{32})")
    m = p.search(info[0][10:])
    if m:
        # creates a direct join link to zoom application
        # another option is to be able to forgo application entirely with
        # link this https://www.zoom.us/wc/join/94165984842?pwd=ME1OemMrdHdUSElGaXdobkN4Z2NzQT09
        # which opens browser version of zoom
        # could have an option but might be confusing to many
        if m.group(1) and m.group(2):
            sendmessage(f"Conference number:")
            sendmessage(f"{m.group(1)}")
            sendmessage(f"Password:")
            sendmessage(f"{m.group(2)}")
    pyautogui.hotkey("winleft", "m")
    try:
        if info[1] != -1:
            webbrowser.open("https://www.zoom.us/j/" + str(info[0]))
            loc = pyautogui.locateCenterOnScreen(
                str(mypath / "images" / "passwordbox.png")
            )
            while loc is None:
                loc = pyautogui.locateCenterOnScreen(
                    str(mypath / "images" / "passwordbox.png")
                )
            time.sleep(1)
            pyautogui.click(
                pyautogui.locateCenterOnScreen(
                    str(mypath / "images" / "closesymbol.png")
                )
            )
            pyautogui.click(loc)
            pyautogui.write(info[1])
        else:
            webbrowser.open(info[0])
            time.sleep(3)
        if joinnewmeeting and pyautogui.locateCenterOnScreen(
            str(mypath / "images" / "anothermeeting.png")
        ):
            yes = pyautogui.locateCenterOnScreen(
                str(mypath / "images" / "meetingyes.png")
            )
            pyautogui.click(yes)
        time.sleep(3)
        joinmeeting = pyautogui.locateCenterOnScreen(
            str(mypath / "images" / "joinmeeting.png")
        )
        if joinmeeting:
            pyautogui.click(joinmeeting)
        time.sleep(3)
        joincomaud = pyautogui.locateCenterOnScreen(
            str(mypath / "images" / "joincomaud.png")
        )
        if joincomaud:
            pyautogui.click(joincomaud)
        time.sleep(2)
        mute = pyautogui.locateCenterOnScreen(str(mypath / "images" / "mute.png"))
        if mute:
            pyautogui.click(mute)
        time.sleep(1)
        fullscreen = pyautogui.locateCenterOnScreen(
            str(mypath / "images" / "fullscreen.png")
        )
        if fullscreen:
            pyautogui.click(fullscreen)
        if platform == "win32":
            winlist = pyautogui.getAllTitles()
            win = pyautogui.getWindowsWithTitle("Zoom Meeting")
            if iszoomopen():
                win[0].maximize()
                sendmessage(
                    f"\U00002705 You have successfully joined your meeting: {info[2].upper()}"
                )
            elif "Waiting for Host" in winlist:
                sendmessage(
                    f"\U000023F3 Waiting for host to start the meeting: {info[2].upper()}"
                )
            else:
                sendmessage(
                    f"\U0000274C ERROR: may have not joined meeting: {info[2].upper()}"
                )
        screenshot()
        print("Finished joining Meeting")
    except IndexError:
        pyautogui.alert("Error data is not correctly entered")
        print("The code/password is not present")


def iskeypresent():
    if "api_key" in globals():
        if str(0) in {api_key, chat_id}:
            return False
        else:
            return True
    else:
        return False


def sendphoto(filepath):
    if iskeypresent():
        logging.info("Sending screenshot to telegram")
        url = f"https://api.telegram.org/bot{api_key}/sendPhoto"
        data = {"chat_id": chat_id}
        files = {"photo": open(filepath, "rb")}
        r = requests.post(url, files=files, data=data)


def sendmessage(mymessage):
    if iskeypresent():
        shortmessage = mymessage[0:10].replace("\n", " ")
        logging.info(f"Sending message: {shortmessage} ...")
        requests.get(
            f"https://api.telegram.org/bot{api_key}/sendMessage",
            params={"chat_id": {chat_id}, "text": {mymessage}},
        )


def makeconfig():
    if not os.path.exists("config.ini"):
        print("Making config file")
        config["Telegram Info"] = {"userid": "0", "api_key": "0"}
        config["Options"] = {"joinnewmeeting": "False"}
        with open("config.ini", "w") as configfile:
            config.write(configfile)
    else:
        config.read("config.ini")


@authenticator
def openzoom(update, context):
    update.message.reply_text("Trying to Open Zoom")
    proc = Popen(r"C:\Users\James\AppData\Roaming\Zoom\bin\zoom.exe")
    time.sleep(7)
    screenshot()


def screenshot():
    if platform == "win32":
        im = pyautogui.screenshot("scrshot.png")
        sendphoto(str(mypath / "scrshot.png"))
    else:
        subprocess.call(["scrot", "-o", "-z", "scrshot.png"])
        sendphoto(str(mypath / "scrshot.png"))


@authenticator
def sendscreenshot(update, context):
    screenshot()


@authenticator
def sendwebcamscr(update, context):
    try:
        cam = cv2.VideoCapture(0)
        frame = cam.read()[1]
        cv2.imwrite("webcam.png", frame)
        sendphoto(str(mypath / "webcam.png"))
    except cv2.error:
        sendmessage("Unable to access Webcam")


def help(update, context):
    """send help message."""
    """
            update.message.reply_text('''There are a few commands with this bot\nIf you type /screen you will be sent a picture of your desktop
    If you type /openzoom it will openzoom''')
    """
    """
    update.message.reply_text('''There are a few commands with this bot\nIf you type /screen you will be sent a picture of your desktop
If you type /sch you will get the message of your schedule sent to you.''')
    """
    sendmessage(
        """There are a few commands with this bot\nIf you type /screen you will be sent a picture of your desktop
If you type /sch you will get the message of your schedule sent to you.\nYou can also send a day example (/sch monday or /sch mon) and get sent your schedule for monday"""
    )
    # update.message.reply_text('''Unauthenticated User''')


@authenticator
def cs(update, context):
    update.message.reply_text("Trying to Open cs accepter")
    pyautogui.press("winleft")
    time.sleep(0.5)
    pyautogui.write("accepter")
    time.sleep(0.5)
    pyautogui.press("enter")
    time.sleep(2)
    d = pyautogui.getWindowsWithTitle("Counter-Strike: Global Offensive")
    if len(d) > 0:
        d[0].maximize()
    screenshot()


@authenticator
def shutit(update, context):
    os.system(f"shutdown /s /t 0")


@authenticator
def workout(update, context):

    workouts = ["*Justin Workout*", "*Coco Workout*", "*8 Minute Abs*"]
    update.message.reply_text(
        f"Today your core workout is:\n{random.choice(workouts)}",
        parse_mode="MarkdownV2",
    )


def checkbreakoutroom():
    loc = pyautogui.locateCenterOnScreen(
        str(mypath / "images" / "join.png"), confidence=0.65
    )
    if loc:
        sendmessage("Trying to join breakout meeting")
        logging.info(loc)
        logging.info("Found join button")
        pyautogui.click(loc)
        time.sleep(7)
        if platform == "win32":
            winlist = pyautogui.getAllTitles()
            window = ""
            win = ""
            strings = ["Room", "Breakout", "breakout", "room"]
            for window in winlist:
                if any(s in window for s in strings):
                    win = pyautogui.getWindowsWithTitle(window)
                    break
            if win != "":
                win[0].activate()
                win[0].maximize()
                sendmessage(
                    f"\U00002705 You have successfully joined your breakoutroom"
                )
                time.sleep(1)
        screenshot()


def logcurtime():
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    logging.debug(f"Current Time = {current_time}")


# need to find better way to check if zoom open


def iszoomopen():
    if platform == "win32":
        strings = ["Zoom Meeting"]
        winlist = pyautogui.getAllTitles()
        for window in winlist:
            if any(s in window for s in strings):
                return True
        return False
    else:
        return True


def checkforquiz():
    buttonlist = list(
        pyautogui.locateAllOnScreen(
            str(mypath / "images" / "quizbutton.png"), grayscale=True
        )
    )
    if len(buttonlist) > 0:
        centerbuttonlist = [pyautogui.center(button) for button in buttonlist]
        ymin = 2000
        highestbutton = ()
        for x in centerbuttonlist:
            if x.y < ymin:
                ymin = x.y
                highestbutton = x
        pyautogui.click(highestbutton)
        sendmessage("Selected poll option")
        screenshot()
        time.sleep(5)
        loc = pyautogui.locateCenterOnScreen(str(mypath / "images" / "submitquiz.png"))
        if loc:
            pyautogui.click(loc)
            sendmessage("Sent poll answer in")
    # findallonscreen for poll button quizbutton.png
    # press the one with the lowest y, should be highest on screen
    pass


def main():
    makeconfig()
    global chat_id
    global api_key
    global joinnewmeeting
    chat_id = config["Telegram Info"]["userid"]
    api_key = config["Telegram Info"]["api_key"]
    joinnewmeeting = config.getboolean("Options", "joinnewmeeting")

    if str(0) in {chat_id, api_key}:
        # pyautogui.alert(text='The chat_id and api_key has not been filled out, please fill out the config.ini as described in the README.\nThen reopen the program',
        #                title='Config Must be Filled out', button='OK')
        print(
            "The chat_id and api_key has not been filled out, if you want to have the telegram bot please fill out the config.ini as described in the README.\nThen reopen the program\n"
        )

    # if config has not been filled out pop up message box and tell user
    # what needs to be filled out
    mybot = ZoomBot()

    if iskeypresent():
        updater = Updater(api_key, use_context=True)

        # Get the dispatcher to register handlers
        dp = updater.dispatcher

        # on different commands - answer in Telegram
        # dp.add_handler(CommandHandler("openzoom", openzoom))
        dp.add_handler(CommandHandler("screen", sendscreenshot))
        dp.add_handler(CommandHandler("cs", cs))
        dp.add_handler(CommandHandler("webcam", sendwebcamscr))
        dp.add_handler(CommandHandler("workout", workout))
        dp.add_handler(CommandHandler("sch", mybot.sendinfo, pass_args=True))
        # dp.add_handler(CommandHandler("shutdown", shutit))
        # on noncommand i.e message - echo the message on Telegram
        dp.add_handler(MessageHandler(Filters.text & ~Filters.command, help))
        updater.start_polling()

    while True:
        if iszoomopen():
            checkbreakoutroom()
            checkforquiz()
        logcurtime()
        schedule.run_pending()
        time.sleep(5)


if __name__ == "__main__":
    main()
